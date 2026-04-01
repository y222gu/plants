"""Physically split data/ into data/train/, data/val/, data/test/ subdirectories.

Moves image directories and annotation files into split-specific subdirectories.
Also creates a spreadsheet recording the split assignments.

Usage:
    python split_data.py              # dry run (print plan, no changes)
    python split_data.py --execute    # actually move files
"""

import argparse
import os
import random
import shutil
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
IMAGE_DIR = DATA_DIR / "image"
ANNOTATION_DIR = DATA_DIR / "annotation"


def discover_samples():
    """Walk image dir to find all annotated samples."""
    annotation_files = {f.name for f in ANNOTATION_DIR.iterdir() if f.is_file()}
    samples = []
    for root, dirs, files in os.walk(IMAGE_DIR):
        if not dirs and files:
            tif_files = [f for f in files if f.lower().endswith((".tif", ".tiff"))]
            if not tif_files:
                continue
            full_path = Path(root)
            rel_path = full_path.relative_to(IMAGE_DIR)
            parts = rel_path.parts
            if len(parts) != 4:
                continue
            species, microscope, experiment, sample_name = parts
            annotation_name = "_".join(parts) + ".txt"
            if annotation_name in annotation_files:
                samples.append({
                    "species": species,
                    "microscope": microscope,
                    "experiment": experiment,
                    "sample_name": sample_name,
                    "uid": f"{species}_{microscope}_{experiment}_{sample_name}",
                    "group_key": f"{species}/{microscope}/{experiment}",
                    "image_dir": full_path,
                    "annotation_path": ANNOTATION_DIR / annotation_name,
                })
    samples.sort(key=lambda s: s["uid"])
    return samples


def compute_split(samples):
    """Compute train/val/test split assignments. Returns uid->split mapping."""
    # --- Test set ---
    test_uids = set()

    # Sorghum/Olympus test experiments
    sorghum_test_exps = {"Exp16", "Exp59", "Exp71", "Exp75", "Exp94", "Exp100", "Exp88", "Exp74"}
    for s in samples:
        if s["species"] == "Sorghum" and s["microscope"] == "Olympus" and s["experiment"] in sorghum_test_exps:
            test_uids.add(s["uid"])

    # Rice/Olympus test experiments
    for s in samples:
        if s["species"] == "Rice" and s["microscope"] == "Olympus" and s["experiment"] in {"Exp10", "Exp11"}:
            test_uids.add(s["uid"])

    # Rice/Zeiss - ALL
    for s in samples:
        if s["species"] == "Rice" and s["microscope"] == "Zeiss":
            test_uids.add(s["uid"])

    # Tomato test - specific samples from spreadsheet
    wb = openpyxl.load_workbook(BASE_DIR / "tomato_samples.xlsx")
    ws = wb.active
    target_rows = [12, 30, 50, 66, 83, 90, 95, 100, 119, 137, 158, 172, 189, 204,
                   223, 231, 260, 286, 312, 323, 344, 387, 405, 450, 466, 471, 480,
                   493, 497, 511]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        if row[0].value in target_rows:
            test_uids.add(row[1].value)

    # Extra 7 Tomato/C10 for test (to reach 10)
    tomato_c10_avail = [s for s in samples
                        if s["species"] == "Tomato" and s["microscope"] == "C10"
                        and s["uid"] not in test_uids]
    rng = random.Random(42)
    rng.shuffle(tomato_c10_avail)
    for s in tomato_c10_avail[:7]:
        test_uids.add(s["uid"])

    # Rice/C10 test experiments
    for s in samples:
        if s["species"] == "Rice" and s["microscope"] == "C10" and s["experiment"] in {"Exp7", "Exp10"}:
            test_uids.add(s["uid"])

    # Sorghum/C10 test experiments
    for s in samples:
        if s["species"] == "Sorghum" and s["microscope"] == "C10" and s["experiment"] in {"Exp3", "Exp4"}:
            test_uids.add(s["uid"])

    # Millet/Olympus test experiments
    for s in samples:
        if s["species"] == "Millet" and s["microscope"] == "Olympus" and s["experiment"] in {"Exp3"}:
            test_uids.add(s["uid"])

    # --- Val: stratified experiment-level, ~10% per species/microscope combo ---
    remaining = [s for s in samples if s["uid"] not in test_uids]

    sm_exps = defaultdict(lambda: defaultdict(list))
    for s in remaining:
        sm_key = f"{s['species']}/{s['microscope']}"
        sm_exps[sm_key][s["group_key"]].append(s)

    val_uids = set()
    rng2 = random.Random(42)

    for sm_key in sorted(sm_exps.keys()):
        exps = sm_exps[sm_key]
        sm_total = sum(len(v) for v in exps.values())
        target = max(1, round(sm_total * 0.1))

        # Pick smallest experiments first to minimize overshoot
        exp_keys = sorted(exps.keys(), key=lambda k: len(exps[k]))

        picked = 0
        for k in exp_keys:
            if picked >= target:
                break
            for s in exps[k]:
                val_uids.add(s["uid"])
            picked += len(exps[k])

    # --- Assign splits ---
    uid_to_split = {}
    for s in samples:
        if s["uid"] in test_uids:
            uid_to_split[s["uid"]] = "test"
        elif s["uid"] in val_uids:
            uid_to_split[s["uid"]] = "val"
        else:
            uid_to_split[s["uid"]] = "train"

    return uid_to_split


def print_summary(samples, uid_to_split):
    """Print split summary table."""
    combos = defaultdict(lambda: {"train": 0, "val": 0, "test": 0})
    for s in samples:
        key = f"{s['species']}/{s['microscope']}"
        combos[key][uid_to_split[s["uid"]]] += 1

    print(f"\n{'Species/Microscope':<25} {'Train':>6} {'Val':>6} {'Test':>6} {'Total':>6}")
    print("-" * 55)
    tt, tv, tte = 0, 0, 0
    for key in sorted(combos.keys()):
        c = combos[key]
        total = c["train"] + c["val"] + c["test"]
        print(f"{key:<25} {c['train']:>6} {c['val']:>6} {c['test']:>6} {total:>6}")
        tt += c["train"]; tv += c["val"]; tte += c["test"]
    print("-" * 55)
    grand = tt + tv + tte
    print(f"{'TOTAL':<25} {tt:>6} {tv:>6} {tte:>6} {grand:>6}")
    print(f"{'%':<25} {100*tt/grand:>5.1f}% {100*tv/grand:>5.1f}% {100*tte/grand:>5.1f}%")


def create_spreadsheet(samples, uid_to_split, output_path):
    """Create xlsx spreadsheet recording the split assignments."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dataset Split"

    # Headers
    headers = ["#", "UID", "Species", "Microscope", "Experiment", "Sample Name", "Split"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="000000"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Split colors
    split_fills = {
        "train": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
        "val": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),     # yellow
        "test": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),    # red
    }

    # Sort samples by species, microscope, experiment, sample_name
    sorted_samples = sorted(samples, key=lambda s: (s["species"], s["microscope"], s["experiment"], s["sample_name"]))

    for i, s in enumerate(sorted_samples, 1):
        split = uid_to_split[s["uid"]]
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=s["uid"])
        ws.cell(row=row, column=3, value=s["species"])
        ws.cell(row=row, column=4, value=s["microscope"])
        ws.cell(row=row, column=5, value=s["experiment"])
        ws.cell(row=row, column=6, value=s["sample_name"])
        split_cell = ws.cell(row=row, column=7, value=split)
        split_cell.fill = split_fills[split]

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    # Add summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Species/Microscope").font = header_font
    ws2.cell(row=1, column=2, value="Train").font = header_font
    ws2.cell(row=1, column=3, value="Val").font = header_font
    ws2.cell(row=1, column=4, value="Test").font = header_font
    ws2.cell(row=1, column=5, value="Total").font = header_font
    for c in range(1, 6):
        ws2.cell(row=1, column=c).fill = header_fill

    combos = defaultdict(lambda: {"train": 0, "val": 0, "test": 0})
    for s in samples:
        key = f"{s['species']}/{s['microscope']}"
        combos[key][uid_to_split[s["uid"]]] += 1

    row = 2
    for key in sorted(combos.keys()):
        c = combos[key]
        total = c["train"] + c["val"] + c["test"]
        ws2.cell(row=row, column=1, value=key)
        ws2.cell(row=row, column=2, value=c["train"])
        ws2.cell(row=row, column=3, value=c["val"])
        ws2.cell(row=row, column=4, value=c["test"])
        ws2.cell(row=row, column=5, value=total)
        row += 1

    # Totals row
    tt = sum(c["train"] for c in combos.values())
    tv = sum(c["val"] for c in combos.values())
    tte = sum(c["test"] for c in combos.values())
    ws2.cell(row=row, column=1, value="TOTAL").font = header_font
    ws2.cell(row=row, column=2, value=tt).font = header_font
    ws2.cell(row=row, column=3, value=tv).font = header_font
    ws2.cell(row=row, column=4, value=tte).font = header_font
    ws2.cell(row=row, column=5, value=tt + tv + tte).font = header_font

    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max_len + 3

    wb.save(output_path)
    print(f"\nSpreadsheet saved to {output_path}")


def move_files(samples, uid_to_split, dry_run=True):
    """Move image directories and annotation files into split subdirectories."""
    moves = []  # (src, dst) pairs

    for s in samples:
        split = uid_to_split[s["uid"]]
        split_image_dir = DATA_DIR / split / "image" / s["species"] / s["microscope"] / s["experiment"] / s["sample_name"]
        split_annotation_dir = DATA_DIR / split / "annotation"
        annotation_name = s["uid"] + ".txt"

        moves.append(("image_dir", s["image_dir"], split_image_dir))
        moves.append(("annotation", s["annotation_path"], split_annotation_dir / annotation_name))

    if dry_run:
        print(f"\nDRY RUN: Would move {len(moves)} items ({len(samples)} samples x 2)")
        # Show a few examples
        for kind, src, dst in moves[:6]:
            print(f"  {kind}: {src.relative_to(BASE_DIR)} -> {dst.relative_to(BASE_DIR)}")
        if len(moves) > 6:
            print(f"  ... and {len(moves) - 6} more")
        return

    # Create split directories
    for split in ("train", "val", "test"):
        (DATA_DIR / split / "image").mkdir(parents=True, exist_ok=True)
        (DATA_DIR / split / "annotation").mkdir(parents=True, exist_ok=True)

    # Move files
    moved = 0
    for kind, src, dst in moves:
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(src), str(dst))
        moved += 1

    print(f"\nMoved {moved} items ({len(samples)} samples)")

    # Clean up empty directories in old image/ and annotation/
    _cleanup_empty_dirs(IMAGE_DIR)
    if ANNOTATION_DIR.exists() and not any(ANNOTATION_DIR.iterdir()):
        ANNOTATION_DIR.rmdir()
        print(f"Removed empty {ANNOTATION_DIR.relative_to(BASE_DIR)}/")
    if IMAGE_DIR.exists() and not any(IMAGE_DIR.iterdir()):
        IMAGE_DIR.rmdir()
        print(f"Removed empty {IMAGE_DIR.relative_to(BASE_DIR)}/")


def _cleanup_empty_dirs(path):
    """Recursively remove empty directories."""
    if not path.is_dir():
        return
    for child in list(path.iterdir()):
        if child.is_dir():
            _cleanup_empty_dirs(child)
    # After cleaning children, check if this dir is now empty
    if path.is_dir() and not any(path.iterdir()):
        path.rmdir()


def main():
    parser = argparse.ArgumentParser(description="Split data into train/val/test directories")
    parser.add_argument("--execute", action="store_true",
                        help="Actually move files (default is dry run)")
    args = parser.parse_args()

    print("Discovering samples...")
    samples = discover_samples()
    print(f"Found {len(samples)} annotated samples")

    print("\nComputing split...")
    uid_to_split = compute_split(samples)

    print_summary(samples, uid_to_split)

    # Create spreadsheet
    spreadsheet_path = BASE_DIR / "dataset_split.xlsx"
    create_spreadsheet(samples, uid_to_split, spreadsheet_path)

    # Move files
    if args.execute:
        print("\nMoving files...")
        move_files(samples, uid_to_split, dry_run=False)
        print("\nDone! Data is now in data/train/, data/val/, data/test/")
    else:
        move_files(samples, uid_to_split, dry_run=True)
        print("\nThis was a DRY RUN. Use --execute to actually move files.")


if __name__ == "__main__":
    main()
